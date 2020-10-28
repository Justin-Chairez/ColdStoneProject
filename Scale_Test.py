import openpyxl
import xlsxwriter
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Opens Excel workbook to export data
wk = openpyxl.load_workbook("Comparison.xlsx")
# Finds the active sheet in the workbook and renames it
sh = wk.active
sh.title = ("Locations and Shake Prices")
# Creates different options for CYO to place prices under

sh["A1"].value = "Location"
sh["B1"].value = "Like it"
sh["C1"].value = "Love it"
sh["D1"].value = "Gotta have it"

# Excel Cell where to start inputting prices
r = 1
c = 2

# Provides Chrome driver for simulation
Path = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(Path)

# Provides URL of Coldstone websites to pull up in driver
with open("URLS.txt") as fp:
    # Reads from a text file the URLs
    urls = fp.read()
    lines = urls.split("\n")
    # Reads each line in the text file
    for line in range(len(lines)):
        values_list = []
        location = ""
        # Moves down one row each time a line is read from the text file
        r += 1
        # Resets the column variable each time a new line is read
        c = 2
        website = "https://" + lines[line]
        driver.get(website)

        # Pulls name of cold stone location and puts it in first column
        menu = driver.find_element_by_class_name("MenuHeader")
        location = menu.find_element_by_tag_name("h1").text
        sh.cell(row=r, column=1).value = location

        try:
            # Clicks the Creations
            creations_link = driver.find_element_by_id("CategoryItem29431")
            creations_link.click()

            # Clicks the Create Your Own Creation link
            create_your_own_link = driver.find_element_by_link_text("Create Your Own Creation")
            create_your_own_link.click()

            # Waits 10secs for pop up menu
            pop_up = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "OptionGroup "))
            )
            # Finds list of prices for 3 sizes
            items = pop_up.find_elements_by_tag_name("li")

            # Iterates through the items
            for item in range(len(items)):
                # Moves past the first item if there are 4 prices listed
                if len(items) == 4 and item == 0:
                    item += 1
                # Adds the price to values list
                else:
                    price = items[item].find_element_by_class_name("option-group-choice-label__attributes").text
                    values_list.append(price)
            # Exports the prices in value list to each respective
            for x in range(len(values_list)):
                sh.cell(row=r, column=c).value = values_list[x]
                c += 1
        # If no creations button is found, moves onto next URL
        except:
            line += 1

driver.quit()
fp.close()
wk.save("Comparison.xlsx")
